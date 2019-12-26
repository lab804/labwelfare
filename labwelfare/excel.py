#
# Copyright (c) Murilo Ijanc' <mbsd@m0x.ru>
#
# Permission to use, copy, modify, and distribute this software for any
# purpose with or without fee is hereby granted, provided that the above
# copyright notice and this permission notice appear in all copies.
#
# THE SOFTWARE IS PROVIDED "AS IS" AND THE AUTHOR DISCLAIMS ALL WARRANTIES
# WITH REGARD TO THIS SOFTWARE INCLUDING ALL IMPLIED WARRANTIES OF
# MERCHANTABILITY AND FITNESS. IN NO EVENT SHALL THE AUTHOR BE LIABLE FOR
# ANY SPECIAL, DIRECT, INDIRECT, OR CONSEQUENTIAL DAMAGES OR ANY DAMAGES
# WHATSOEVER RESULTING FROM LOSS OF USE, DATA OR PROFITS, WHETHER IN AN
# ACTION OF CONTRACT, NEGLIGENCE OR OTHER TORTIOUS ACTION, ARISING OUT OF
# OR IN CONNECTION WITH THE USE OR PERFORMANCE OF THIS SOFTWARE.
#
"""Module manipulate excel file.

The module facilitates tasks involving excel format files. Tasks like entering
data, get data, convert data.
"""
import logging
import re

from openpyxl import load_workbook

LOGGER = logging.getLogger(__name__)


def get_ws(wb, ws_index):
    """Getting working sheet.

    Args:
        ws (openpyxl.worksheet.workbook.Workbook): workbook.
        ws_index (int): index work sheet.

    Returns:
        ws (openpyxl.worksheet.worksheet.Worksheet): working sheet.

    Raises:
        ValueError: Not found worksheet index.
    """
    # dimension work sheet
    dimension = None
    # name eg. Sheet 1
    ws_name = None
    ws = None

    try:
        ws_name = wb.sheetnames[ws_index]
    except IndexError:
        LOGGER.exception('Not found work sheet by index: {}'.format(ws_index))
        raise ValueError('Not found work sheet by index: {}'.format(ws_index))

    # get worksheet if None set active default
    ws = wb[ws_name]
    if ws is None:
        ws = wb.active
    LOGGER.info('getting working sheet: %s' % (ws_name))

    # checking dimension of file
    dimension = ws.calculate_dimension()
    if dimension is None or dimension == 'A1:A1':
        ws.reset_dimensions()

    return ws


def read(filename, ws_index=0):
    """Open and read excel.

    Args:
        filename (string): file path.

    Returns:
        float: heat load index value

    Raises:
        OSError: File not found.
    """
    wb = load_workbook(filename=filename, read_only=True)
    return wb


def find_col_by_re(patterns,
                   ws,
                   row_limit=None,
                   col_limit=None,
                   multiple=False):
    """Find column by regex pattern.

    Args:
        pattern (string): pattern.
        ws (openpyxl.worksheet.worksheet.Worksheet): worksheet.
        row_limit (int): limit to row find, default all rows.
        col_limit (int): limit to col find, default all columns.

    Returns:
        tuple: containing the row and column of the pattern instance.
    """
    # TODO: check limit of row and column
    first_matches = []
    for row in ws.iter_rows(max_row=row_limit, max_col=col_limit):
        # rcell equal cell read only
        for rcell in row:
            if rcell.value and re.match(patterns[0], rcell.value):
                LOGGER.debug("Found pattern: {} on row: {}, col: {}".format(
                    patterns[0], rcell.row, rcell.column))
                if multiple:
                    first_matches.append((rcell.row, rcell.column))
                else:
                    return (rcell.row, rcell.column)
        if len(first_matches) > 0:
            break
    # "multiple" in case the pattern is on two lines ex. connect
    # worksheet column 3: Row 2 == Temp. and Line 3 == Ext.
    # Below is only true when the column of the previous row is the
    # same as the current row the counter has 2 and obviously the
    # multiple argument is true.
    for row, col in first_matches:
        rcell = ws.cell(row + 1, col)
        if rcell.value and re.match(patterns[1], rcell.value):
            return (rcell.row, rcell.column)

    return (None, None)
